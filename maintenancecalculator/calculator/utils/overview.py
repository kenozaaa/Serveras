import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle

def create_overview_sheet(workbook, results_df):
    # Ensure that the 'Total Fees' column is numeric
    results_df['Total Fees'] = pd.to_numeric(results_df['Total Fees'], errors='coerce')

    # Total fees per country
    if 'Publication Country' not in results_df.columns:
        raise ValueError("Missing 'Publication Country' column in the results sheet")

    # Identify year columns
    year_columns = [col for col in results_df.columns if col.isdigit()]
    if not year_columns:
        raise ValueError("No year columns found (digit-only column names expected)")

    # Total fees per year (ensure the year columns are numeric)
    total_fees_per_country = results_df.groupby('Publication Country')['Total Fees'].sum().reset_index()

    # Ensure that the year columns are numeric before summing
    results_df[year_columns] = results_df[year_columns].apply(pd.to_numeric, errors='coerce')
    total_fees_per_year = results_df[year_columns].sum().reset_index()
    total_fees_per_year.columns = ['Year', 'Maintenance Cost ($)']

    # Create a new sheet in the workbook for overview
    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
    else:
        overview_sheet = workbook.create_sheet(title='Overview')

    # Add data to the overview sheet
    overview_sheet.append(['Publication Country', 'Total Fees'])
    for row in dataframe_to_rows(total_fees_per_country, index=False, header=False):
        overview_sheet.append(row)

    overview_sheet.append([])  # Add some space
    overview_sheet.append(['Year', 'Maintenance Cost ($)'])
    for row in dataframe_to_rows(total_fees_per_year, index=False, header=False):
        overview_sheet.append(row)

    # Format headers (same as before)
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    for col in overview_sheet.iter_rows(min_row=1, max_row=1):
        for cell in col:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill

    # Adjust column widths
    for col in range(1, overview_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        overview_sheet.column_dimensions[col_letter].width = 15



def format_dates_and_currency(workbook):
    """
    Apply date and currency formatting to the in-memory workbook.
    
    Parameters:
    - workbook: An openpyxl Workbook object (in-memory).
    """
    main_sheet = workbook.active  # Assuming the main sheet is the first one

    # Define date and currency styles
    date_style = NamedStyle(name='short_date', number_format='MM/DD/YYYY')
    currency_style = NamedStyle(name='currency', number_format='$#,##0.00')
    
    # Apply date formatting to specific columns
    date_columns = ['C', 'D', 'E', 'F']  # Assuming these columns contain dates
    for col in date_columns:
        for cell in main_sheet[col][1:]:  # Skip header row
            cell.style = date_style

    # Apply currency formatting starting from column K (11th column)
    for col in range(11, main_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        for cell in main_sheet[col_letter][1:]:
            cell.style = currency_style

    # Apply currency formatting to the Overview sheet, if it exists
    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
        for cell in overview_sheet['B'][1:]:  # Assuming column B has currency
            cell.style = currency_style